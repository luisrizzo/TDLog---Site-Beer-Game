from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Alignment
import datetime
import json
import os
import shutil
import numpy as np
import pdb
from pathlib import Path
import Environment as Env
import Demand as Demand
import Trainer as Tr

class Saver:
	"""
	Class to save results of the AI
	Parameters : 
		path :
		results_folder_name :
		excel_file_name :
		...
	"""
	def __init__(self, path):
		"""
		Initialize the save instance 
		Args :
			path : the path where to save the results
		"""
		self.absolute_path = path #'/Users/hamzakabbaj/Dropbox/BeerGame/'
		self.results_folder_name = "results"
		self.results_file_name = "beerGameResults.xlsx"

		self.results_folder_path = self.absolute_path + self.results_folder_name
		self.results_file_path = self.absolute_path + self.results_file_name

		self.columns_list = ['ID', # Test ID
							#### GAME PARAMETERS
							'CLT_DEMAND',  # Client Demand
							"CP_AGENT" , # Comparing agent
							'AGENTS', # List of agents labels 
							'USE_BO', # Using backorders
							'T',  # Number of game periods
							'LDT',  # agents Lead times 
							'HC',  # Holding costs
							'SC',  # Shortage Cost
							'SR',  # Service rate
							'IIL', # Initial Inventory Level

							#### AI PARAMETERS
							'ACTIONS',  # list of possible actions
							'm', # number of states observed by the AI 
							'AI_DN',  # AI Deep Network
							'N_ITER', # Number of iterations 
							"TIME_PERF", # Time / 100 iteration

							#### RESULTS
							 
							'AVG_SUM_DEMAND', # Sum of demand
							'AI_AVG_CUM_COSTS', # AI Average Cumulated Costs
							'CP_AVG_CUM_COSTS', # CP Average Cumulated Costs
							#### AI RESULTS
							'AI_AVG_CR', # AI Average Coverage Rate
							"AI_AVG_BR", # AI Average Breakdown Rate
							"AI_AVG_SR", # AI Average Service rate
							#### CP RESULTS
							'CP_AVG_CR', # CP Average Coverage Rate
							"CP_AVG_BR", # CP Average Breakdown Rate
							"CP_AVG_SR", # CP Average Service rate
							"TEST_DATE"]

	def create(self):
		"""
		Creates file and folders where to save the results
		"""
		# Creating folders where results are saved
		if not os.path.exists(Path(self.results_folder_path)):
			os.makedirs(Path(self.results_folder_path))

		# Creating Excel file to save numeric results
		if not os.path.isfile(Path(self.results_file_path)):
			wb = Workbook()

			ws = wb.active
			ws.append(self.columns_list)

			wb.save(Path(self.results_file_path))

	def get_test_id(self):
		"""
		Returns the test id based on the existing folders
		"""
		# initialize the test id
		test_id = 1
		while os.path.exists(Path(self.results_folder_path+'/'+str(test_id))):
			test_id += 1

		return test_id

	def save(self, trainer):
		"""
		Saves the results of trainer in a file
		Args ; 
			trainer :
		"""
		# create results folders and files
		self.create()

		results_dict = {key: None for key in self.columns_list}

		# Launches a comparison before saving results
		trainer.comparator.launch_comparison(50)

		lead_times = []
		for agent in trainer.agents:
			lead_times.append(agent.lead_time.display())

		test_id = self.get_test_id()

		results_dict['ID'] = test_id
		#### GAME PARAMETERS
		results_dict['CLT_DEMAND'] = trainer.env.params['client_demand'].display()
		results_dict['CP_AGENT'] = trainer.params['comparison_agent'].label
		results_dict['AGENTS'] = trainer.get_agents_labels()
		results_dict['USE_BO'] = trainer.params['use_backorders']
		results_dict['T'] = trainer.params['number_periods']
		results_dict['LDT'] = str(lead_times)
		results_dict['HC'] = trainer.params['holding_cost']
		results_dict['SC'] = trainer.params['shortage_cost']
		results_dict['SR'] = trainer.params['TS']
		results_dict['IIL'] = trainer.params['initial_inventory']

		#### AI PARAMETERS

		results_dict['ACTIONS'] = str("(min = "+str(np.min(trainer.params['AI_possible_actions']))+" , max = "+str(np.max(trainer.params['AI_possible_actions']))+")")
		results_dict['m'] = trainer.params['m']
		results_dict['AI_DN'] = str(trainer.params['AI_DN'])
		results_dict['N_ITER'] = trainer.train_iter
		results_dict["TIME_PERF"] = round(trainer.time_per_iteration * 100,2)

		#### RESULTS
		results_dict['AVG_SUM_DEMAND'] = trainer.comparator.AI_performance['sum_demand']
		results_dict['AI_AVG_CUM_COSTS'] = trainer.comparator.AI_performance['costs']
		results_dict['CP_AVG_CUM_COSTS'] = trainer.comparator.CP_performance['costs']
		
		#### AI RESULTS
		results_dict['AI_AVG_CR'] = trainer.comparator.AI_performance['coverage_rate']
		results_dict["AI_AVG_BR"] = trainer.comparator.AI_performance['breakdown_rate']
		results_dict["AI_AVG_SR"] = trainer.comparator.AI_performance['service_rate']
		#### CP RESULTS
		results_dict['CP_AVG_CR'] = trainer.comparator.CP_performance['coverage_rate']
		results_dict["CP_AVG_BR"] = trainer.comparator.CP_performance['breakdown_rate']
		results_dict["CP_AVG_SR"] = trainer.comparator.CP_performance['service_rate']
		#### DATE 
		results_dict["TEST_DATE"] = datetime.datetime.now().replace(second=0, microsecond=0)

		# Convert result dictionnary into a list
		results = list(results_dict.values())

		# Load the excel workbook
		wb = load_workbook(Path(self.results_file_path))
		ws = wb.active

		i = 1
		while ws['A'+str(i)].value != None:
			i += 1
		for j, result in enumerate(results):
			cell = ws.cell(column = j+1, row=i, value=result)
			cell.alignment = Alignment(horizontal='center')
		
		wb.save(Path(self.results_file_path))
		

		# CREATE A NEW FOLDER IN FOLDER RESULTS
		#os.makedirs(Path(self.results_folder_path+'/'+str(test_id)))

		# Save AI Model
		#trainer.agents[trainer.AI_index].save_model(os.fspath(Path(self.results_folder_path+"/"+str(test_id)+"/tfmodel/model.ckpt")))
		
		# TO DO !!!
		# Save Trainer class
		#trainer.save_json(self.results_folder_path+"/"+str(test_id)+'/')

		# Save Figures :
		#comparator.histogram_fig.savefig(os.fspath(Path(self.results_folder_path+"/"+str(test_id)+"/histogram_fig.png")))
		#comparator.game_fig.savefig(os.fspath(Path(self.results_folder_path+"/"+str(test_id)+"/game_fig.png")))
		#trainer.figure_displayer.train_fig.fig.savefig(os.fspath(Path(self.results_folder_path+"/"+str(test_id)+"/train_fig.png")))

	def load_demand(self, id):
		with open(Path(self.path+"results/"+str(id)+'/demand.json'), 'r') as input_file:
			demand_json = json.load(input_file)

		if demand_json['label'] == "Gaussian":
			demand = Demand.Gaussian_Demand(
					demand_json['Mu'],
					demand_json['Sigma'],
					demand_json['Min'],
					demand_json['Max']
				)

		elif demand_json['label'] == "Uniform":
			demand = Demand.Uniform_Demand(
					demand_json['Min'],
					demand_json['Max'],
					demand_json['Step'],
				)

		else:
			assert False, "Couldn't load demand"

		return demand

	def load_env(self, id):
		with open(Path(self.path+"results/"+str(id)+'/env.json'), 'r') as input_file:
			env_json = json.load(input_file)

		cliend_demand = self.load_demand(id)
		env = Env.Environment(cliend_demand, 
								agents = [], 
								params = env_json['params'],
								state_features =env_json['state_features'])

		env.NA = env_json['NA']
		env.max_lead_time = env_json['max_lead_time']
		env.nb_state_features = env_json['nb_state_features']

		return env

	def load(self, id):
		with open(Path(self.path+"results/"+str(id)+'/trainer.json'), 'r') as input_file:
			trainer_json = json.load(input_file)

		
		#env = self.load_env(id)

		cliend_demand = self.load_demand(id)

		trainer = Tr.Trainer(client_demand = cliend_demand,
							 AI_possible_actions = trainer_json['AI_possible_actions'],
							 agents_labels = trainer_json['agents_labels'],
							 params = trainer_json['params'],
							 state_features = trainer_json['state_features'])
	
		trainer.AI_index = trainer_json['AI_index']
		trainer.train_iter = trainer_json['train_iter']
		trainer.time_per_iteration = trainer_json['time_per_iteration']
		trainer.AI_agent_costs = trainer_json['AI_agent_costs']
		trainer.AI_agent_smooth_costs = trainer_json['AI_agent_smooth_costs']
		trainer.AI_agent_best_quartil_costs = trainer_json['AI_agent_best_quartil_costs']
		trainer.AI_agent_worst_quartil_costs = trainer_json['AI_agent_worst_quartil_costs']

		trainer.create_agents()
		res_path = self.path+"results/"+str(id)+"/tfmodel/model.ckpt"
		p  = Path(res_path)

		trainer.agents[trainer.AI_index].load(p)
		return trainer

	def load_DQN_agent(self, id, DQN_agent):
		res_path = self.path+"results/"+str(id)+"/tfmodel/model.ckpt"
		p  = Path(res_path)
		DQN_agent.load(p)

	def clean_results_folder(self):
		"""
		Deletes folders that has no id in excel file
		"""
		if os.path.exists(Path(self.results_file_path)):
			wb = load_workbook(Path(self.results_file_path))
			ws = wb.active

			trainers_ids = []

			i = 1
			while ws['A'+str(i)].value != None:
				trainers_ids.append(ws['A'+str(i)].value)
				i += 1


			trainers_dirs = os.listdir(Path(self.results_folder_path))

			for tr_dir in trainers_dirs:
				if tr_dir.isdigit() and (int(tr_dir) not in trainers_ids):
					print(tr_dir,"deleted !! ")
					shutil.rmtree(Path(self.results_folder_path+'/'+tr_dir))

		

